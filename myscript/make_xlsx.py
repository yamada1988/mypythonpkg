import openpyxl
import numpy as np
from openpyxl.styles import Font
import datetime
import os

As = [chr(ord('A') + i) for i in range(26)]


dirname = os.getcwd().split('/')[-1]
now = datetime.datetime.now()
date = "{0:%Y-%m-%d-%H}".format(now)
ofname = dirname + '-' + date + '.xlsx'  

wb = openpyxl.Workbook()
#nms = [('01', '48'), ('02', '24'), ('03', '16'), ('04', '12'), ('06', '08'), ('08', '06'), ('12', '04'), ('24', '02'), ('48', '01')]
nms = [('01', '48'), ('02', '24'), ('04', '12'), ('06', '08'), ('08', '06'), ('12', '04'), ('24', '02'), ('48', '01')]
inds = [1] + list(range(5,100+1,5))
inds=range(55,100+1,5)
N = len(inds)
Nk = 48

for nm in nms:
    print(nm[0])
    titlename = 'conf_'+ nm[0]+'_dmus_'+nm[1]
    print(titlename)
    ws = wb.create_sheet(title=titlename) 
    ws["B1"] = "mu (kcal/mol)"
    ws["C1"] = "error (kcal/mol)"
    ws["C1"] = "95%err"  
 
    n = int(nm[0])
    m = int(nm[1])
    mus = [[] for i in range(N)]
    errors = [[] for i in range(N)]
    mesherrs = [[] for i in range(N)]
    for ii,i in enumerate(inds):
        ws["A{0:d}".format(ii+2)] = "{0:03d}".format(i)
        for k in range(1,m+1):
            ws.cell(row=1, column=k+4, value='sln{0:d}'.format(k))
            fname = 'MELT_{0:03d}/conf_{1:02d}_dmus_{2:02d}/slvfe_{3:02d}/slvfe.tt'.format(i, n, m, k)
            with open(fname, 'rt') as f:
                lines = f.readlines()
            l3 = lines[-3].split()
            l1 = lines[-1].split()
            print(l3)
            mu = float(l3[1])
            error = float(l3[2])
            mesherr = float(l1[4])
 
            ws.cell(row=ii+2, column=k+4, value=mu)
    
            mus[ii].append(mu)
            errors[ii].append(error)
            mesherrs[ii].append(mesherr)
    
        mus[ii] = np.array(mus[ii])
        errors[ii] = np.array(errors[ii])
        mesherrs[ii] = np.array(mesherrs[ii])
        err95 = 2.0*np.std(mus[ii])/np.sqrt(m)
        print(i, np.mean(mus[ii]), np.std(mus[ii]), np.mean(errors[ii]))
    
        #ws.cell(row=ii+2, column=2, value='AVERAGE(D{0:d}:)'.format(ii+2))
        #ws.cell(row=ii+2, column=3, value='STDEV.S(D{0:d}:'.format(ii+2))
        ws.cell(row=ii+2, column=2, value=np.mean(mus[ii]))
        ws.cell(row=ii+2, column=3, value=np.std(mus[ii]))
        ws.cell(row=ii+2, column=4, value=err95) 
    
    
    font = Font(name='Arial', size=14)
    for row in ws:
        for cell in row:
            ws[cell.coordinate].font = font
    wb.save(ofname)
